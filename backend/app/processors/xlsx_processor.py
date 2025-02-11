import io

import openpyxl

from app.processors.document_processor import DocumentProcessor


class XLSXFileProcessor(DocumentProcessor):
    def extract_text(self, file_content) -> str:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        text = ''
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += ' '.join([str(cell) for cell in row]) + '\n'
        return text
